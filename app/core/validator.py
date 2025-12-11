# app/core/validator.py
from openpyxl import load_workbook

class Validator:
    @staticmethod
    def validate_file(file_obj, required_sheets=None):
        if required_sheets is None:
            required_sheets = ['Reguler', 'Poleks']
        try:
            wb = load_workbook(file_obj, read_only=True)
            missing = [s for s in required_sheets if s not in wb.sheetnames]
            if missing:
                return False, f"Missing sheets: {', '.join(missing)}"
            return True, None
        except Exception as e:
            return False, str(e)

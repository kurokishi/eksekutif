from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import io
from datetime import datetime, timedelta
import pandas as pd


class ExcelWriter:

    def __init__(self, config):
        self.config = config

        # warna slot
        self.fill_r = PatternFill(start_color="00FF00", fill_type="solid")
        self.fill_e = PatternFill(start_color="0000FF", fill_type="solid")
        self.fill_over = PatternFill(start_color="FF0000", fill_type="solid")

        # warna untuk peta konflik
        self.fill_conflict_normal = PatternFill(start_color="FFFF00", fill_type="solid")  # kuning
        self.fill_conflict_hard = PatternFill(start_color="FF0000", fill_type="solid")    # merah

        # border header
        self.border_header = Border(bottom=Side(border_style="thick"))

    # ======================================================================
    # HELPER – gabung range jam
    # ======================================================================
    def _combine_ranges(self, slots, interval):
        if not slots:
            return []
        ts = sorted(datetime.strptime(t, "%H:%M") for t in slots)
        ranges = []
        start = ts[0]
        end = start + timedelta(minutes=interval)

        for t in ts[1:]:
            if t == end:
                end = t + timedelta(minutes=interval)
            else:
                ranges.append((start, end))
                start = t
                end = start + timedelta(minutes=interval)
        ranges.append((start, end))
        return ranges

    def _format_range(self, a, b):
        return f"{a.strftime('%H.%M')}–{b.strftime('%H.%M')}"

    # ======================================================================
    # FEATURE 1 — Peak Hour Analysis
    # ======================================================================
    def _create_peak_hour(self, wb, df, slot_str):
        if "Peak Hour Analysis" in wb.sheetnames:
            del wb["Peak Hour Analysis"]
        ws = wb.create_sheet("Peak Hour Analysis")

        ws.append(["HARI", "SLOT", "JUMLAH", "KATEGORI"])

        for hari, g in df.groupby("HARI"):
            counts = {s: 0 for s in slot_str}
            for _, row in g.iterrows():
                for slot in slot_str:
                    if row.get(slot) in ["R", "E"]:
                        counts[slot] += 1

            max_v = max(counts.values())
            kategori = "High Load" if max_v >= 10 else "Medium" if max_v >= 5 else "Low"

            for slot, v in counts.items():
                if v == max_v:
                    ws.append([hari, slot, v, kategori])

    # ======================================================================
    # FEATURE 2 — Conflict Checking Dokter (textual)
    # ======================================================================
    def _create_conflict_doctor(self, wb, df, slot_str):
        if "Conflict Dokter" in wb.sheetnames:
            del wb["Conflict Dokter"]

        ws = wb.create_sheet("Conflict Dokter")
        ws.append(["DOKTER", "HARI", "SLOT", "KONFLIK"])

        for (dokter, hari), g in df.groupby(["DOKTER", "HARI"]):
            for slot in slot_str:
                vals = g[slot].unique()

                # Konflik poli berbeda
                if len(vals) > 1 and any(v in ["R", "E"] for v in vals):
                    ws.append([dokter, hari, slot, "Dokter memiliki 2 poli berbeda di jam yang sama"])

                # Konflik R & E bersamaan
                if "R" in vals and "E" in vals:
                    ws.append([dokter, hari, slot, "Bentrok Reguler & Poleks"])

    # ======================================================================
    # FEATURE 3 — VISUAL CONFLICT MAP
    # ======================================================================
    def _create_conflict_map(self, wb, df, slot_str):
        if "Peta Konflik Dokter" in wb.sheetnames:
            del wb["Peta Konflik Dokter"]

        ws = wb.create_sheet("Peta Konflik Dokter")

        # Header
        doctors = sorted(df["DOKTER"].unique())
        ws.append(["SLOT"] + doctors)

        for slot in slot_str:
            row = [slot] + ["" for _ in doctors]
            ws.append(row)

        # mapping dokter ke kolom
        doc_index = {doc: i + 2 for i, doc in enumerate(doctors)}

        # isi warna konflik
        for (dokter, hari), g in df.groupby(["DOKTER", "HARI"]):
            col = doc_index[dokter]

            for row_idx, slot in enumerate(slot_str, start=2):
                vals = g[slot].unique()

                if len(vals) > 1 and any(v in ["R", "E"] for v in vals):
                    ws.cell(row=row_idx, column=col).fill = self.fill_conflict_normal
                if "R" in vals and "E" in vals:
                    ws.cell(row=row_idx, column=col).fill = self.fill_conflict_hard

    # ======================================================================
    # FEATURE 4 — Rekap Layanan (range)
    # ======================================================================
    def _create_rekap_layanan(self, wb, df, slot_str):
        if "Rekap Layanan" in wb.sheetnames:
            del wb["Rekap Layanan"]

        ws = wb.create_sheet("Rekap Layanan")
        ws.append(["POLI", "HARI", "DOKTER", "JENIS", "JAM LAYANAN"])

        interval = self.config.interval_minutes

        for (poli, hari, dokter), g in df.groupby(["POLI ASAL", "HARI", "DOKTER"]):
            R = [s for s in slot_str if g.iloc[0].get(s) == "R"]
            E = [s for s in slot_str if g.iloc[0].get(s) == "E"]

            for a, b in self._combine_ranges(R, interval):
                ws.append([poli, hari, dokter, "Reguler", self._format_range(a, b)])

            for a, b in self._combine_ranges(E, interval):
                ws.append([poli, hari, dokter, "Poleks", self._format_range(a, b)])

    # ======================================================================
    # FEATURE 5 — Rekap POLI
    # ======================================================================
    def _create_rekap_poli(self, wb, df, slot_str):
        if "Rekap Poli" in wb.sheetnames:
            del wb["Rekap Poli"]

        ws = wb.create_sheet("Rekap Poli")
        ws.append(["POLI", "HARI", "TOTAL REG", "TOTAL POLEKS", "TOTAL JAM"])

        interval = self.config.interval_minutes

        for (poli, hari), g in df.groupby(["POLI ASAL", "HARI"]):
            tot_r = sum((g.iloc[0].get(s) == "R") * interval/60 for s in slot_str)
            tot_e = sum((g.iloc[0].get(s) == "E") * interval/60 for s in slot_str)
            ws.append([poli, hari, round(tot_r,2), round(tot_e,2), round(tot_r+tot_e,2)])

    # ======================================================================
    # FEATURE 6 — Rekap Dokter dengan PENGGABUNGAN SHIFT
    # ======================================================================
    def _create_rekap_dokter(self, wb, df, slot_str):
        if "Rekap Dokter" in wb.sheetnames:
            del wb["Rekap Dokter"]

        ws = wb.create_sheet("Rekap Dokter")
        ws.append(["DOKTER", "HARI", "SHIFT", "TOTAL JAM"])

        interval = self.config.interval_minutes

        for (dokter, hari), g in df.groupby(["DOKTER", "HARI"]):

            # kumpulkan slot aktif (R/E)
            active_slots = [s for s in slot_str if g.iloc[0].get(s) in ["R", "E"]]

            # gabungkan shift otomatis
            merged = self._combine_ranges(active_slots, interval)

            for a, b in merged:
                dur = (b - a).seconds / 3600
                ws.append([
                    dokter,
                    hari,
                    self._format_range(a, b),
                    round(dur, 2)
                ])

    # ======================================================================
    # FEATURE 7 — Grafik Beban POLI
    # ======================================================================
    def _create_grafik_poli(self, wb):
        if "Grafik Beban Poli" in wb.sheetnames:
            del wb["Grafik Beban Poli"]

        ws = wb.create_sheet("Grafik Beban Poli")
        ws["A1"] = "Grafik Beban Poli per Minggu"

        rp = wb["Rekap Poli"]

        table = {}
        for row in rp.iter_rows(min_row=2, values_only=True):
            table[row[0]] = table.get(row[0], 0) + row[4]

        ws.append(["POLI", "TOTAL JAM"])
        for k, v in table.items():
            ws.append([k, v])

        chart = BarChart()
        chart.title = "Beban Poli"

        data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        chart.add_data(data)
        chart.set_categories(cats)
        ws.add_chart(chart, "E5")

    # ======================================================================
    # SHEET UTAMA: Jadwal (tanpa border antar hari)
    # ======================================================================
    def write(self, source_file, df, slot_str):
        wb = load_workbook(source_file)

        # clean old
        if "Jadwal" in wb.sheetnames:
            del wb["Jadwal"]
        ws = wb.create_sheet("Jadwal")

        headers = ["POLI ASAL", "JENIS POLI", "HARI", "DOKTER"] + slot_str
        ws.append(headers)

        for _, r in df.iterrows():
            ws.append([r.get(h, "") for h in headers])

        # pewarnaan
        self.apply_styles(ws, df, slot_str)

        # semua fitur rekap
        self._create_rekap_layanan(wb, df, slot_str)
        self._create_rekap_poli(wb, df, slot_str)
        self._create_rekap_dokter(wb, df, slot_str)
        self._create_peak_hour(wb, df, slot_str)
        self._create_conflict_doctor(wb, df, slot_str)
        self._create_conflict_map(wb, df, slot_str)
        self._create_grafik_poli(wb)

        # finishing
        self._auto_width_all_sheets(wb)
        self._style_headers_all(wb)
        self._freeze_headers_all(wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ======================================================================
    # Pewarnaan slot (no border per day)
    # ======================================================================
    def apply_styles(self, ws, df, slot_str):
        counter = {h: {s: 0 for s in slot_str} for h in df["HARI"].unique()}
        excel_row = 2

        for rec in df.to_dict("records"):
            hari = rec["HARI"]

            for idx, slot in enumerate(slot_str):
                val = rec.get(slot, "")
                col_idx = 5 + idx
                cell = ws.cell(row=excel_row, column=col_idx)

                if val == "R":
                    cell.fill = self.fill_r
                elif val == "E":
                    counter[hari][slot] += 1
                    if counter[hari][slot] > self.config.max_poleks_per_slot:
                        cell.fill = self.fill_over
                    else:
                        cell.fill = self.fill_e

            excel_row += 1

    # ======================================================================
    # STYLING
    # ======================================================================
    def _auto_width_all_sheets(self, wb):
        for ws in wb.worksheets:
            for col in ws.columns:
                width = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = width + 2

    def _style_headers_all(self, wb):
        for ws in wb.worksheets:
            for c in ws[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center")
                c.border = self.border_header

    def _freeze_headers_all(self, wb):
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
